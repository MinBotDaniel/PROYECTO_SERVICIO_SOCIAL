import os
import sqlite3
import openpyxl

# --- CONFIGURACIÓN ---
CARPETA_EXCEL = "C:/PROYECTO_SERVICIO_SOCIAL/ArchivosExcel"  # <--- RECUERDA CAMBIAR ESTO
NOMBRE_DB = "tienda.db"

def inicializar_db():
    conn = sqlite3.connect(NOMBRE_DB)
    cursor = conn.cursor()
    
    # 1. Tabla Clientes
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT,
            domicilio TEXT,
            telefono TEXT,
            archivo_origen TEXT
        )
    ''')

    # 2. Tabla Ventas (Ya sin el folio_vale)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            fecha TEXT,
            quincenas INTEGER,
            total REAL,
            adeudo REAL,
            estado TEXT,
            FOREIGN KEY (cliente_id) REFERENCES clientes (id)
        )
    ''')

    # 3. Tabla Pagos (Abonos)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pagos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            venta_id INTEGER,
            numero_pago INTEGER,
            fecha TEXT,
            monto REAL,
            FOREIGN KEY (venta_id) REFERENCES ventas (id)
        )
    ''')
    conn.commit()
    return conn

def migrar_archivos():
    conn = inicializar_db()
    cursor = conn.cursor()
    
    archivos = [f for f in os.listdir(CARPETA_EXCEL) if f.endswith('.xlsx') and not f.startswith('~')]
    total_archivos = len(archivos)
    
    print(f"Encontrados {total_archivos} archivos. Iniciando migración...")

    for i, archivo in enumerate(archivos):
        ruta_completa = os.path.join(CARPETA_EXCEL, archivo)
        
        # Evitar duplicados
        cursor.execute("SELECT id FROM clientes WHERE archivo_origen = ?", (archivo,))
        if cursor.fetchone():
            print(f"[{i+1}/{total_archivos}] Saltando {archivo} (Ya existe)")
            continue
            
        try:
            wb = openpyxl.load_workbook(ruta_completa, data_only=True)
            hoja = wb.active
            
            nombre, domicilio, telefono = "Desconocido", "", ""
            
            # --- A. EXTRAER CLIENTE (CON COORDENADAS EXACTAS) ---
            nombre, domicilio, telefono = "Desconocido", "", ""
            
            fila_titulos = 2
            fila_datos = 4
            
            # Buscamos en TODA la fila 2 (hasta la columna 100 por si es muy ancho)
            for col in range(1, 100):
                celda_titulo = hoja.cell(row=fila_titulos, column=col).value
                
                if isinstance(celda_titulo, str):
                    texto = celda_titulo.strip().upper()
                    
                    if "NOMBRE" in texto:
                        # Bajamos a la fila 4 y agarramos el texto
                        for c in range(col, col+5):
                            val = hoja.cell(row=fila_datos, column=c).value
                            if val is not None and str(val).strip() != "":
                                nombre = str(val).strip()
                                break
                                
                    elif "DOMICILIO" in texto:
                        for c in range(col, col+5):
                            val = hoja.cell(row=fila_datos, column=c).value
                            if val is not None and str(val).strip() != "":
                                domicilio = str(val).strip()
                                break
                                
                    elif "TELEFONO" in texto or "TEL" in texto:
                        for c in range(col, col+5):
                            val = hoja.cell(row=fila_datos, column=c).value
                            if val is not None and str(val).strip() != "":
                                telefono = str(val).strip()
                                break
            # -------------------------------------------------------

            cursor.execute("INSERT INTO clientes (nombre, domicilio, telefono, archivo_origen) VALUES (?, ?, ?, ?)", 
                           (str(nombre), str(domicilio), str(telefono), archivo))
            cliente_id = cursor.lastrowid

            # --- B. EXTRAER VALES Y SUS PAGOS ---
            for fila in range(1, hoja.max_row + 1):
                for col in range(1, 15):
                    celda = hoja.cell(row=fila, column=col).value
                    
                    if isinstance(celda, str) and celda.strip().upper() == "VALE":
                        # Datos de la venta (Omitimos el folio)
                        fecha_vale = hoja.cell(row=fila+1, column=col+1).value
                        quincenas = hoja.cell(row=fila+1, column=col+2).value
                        
                        total = 0.0
                        adeudo = 0.0
                        for col_derecha in range(col+3, hoja.max_column + 1):
                            titulo_derecha = hoja.cell(row=fila, column=col_derecha).value
                            if isinstance(titulo_derecha, str):
                                titulo = titulo_derecha.strip().upper()
                                if titulo == "TOTAL":
                                    val = hoja.cell(row=fila+1, column=col_derecha).value
                                    total = float(val) if val is not None else 0.0
                                elif titulo == "ADEUDO":
                                    val = hoja.cell(row=fila+1, column=col_derecha).value
                                    adeudo = float(val) if val is not None else 0.0
                        
                        estado = "Pendiente" if adeudo > 0 else "Pagado"

                        # Insertar Venta (sin folio_vale)
                        cursor.execute('''INSERT INTO ventas 
                                          (cliente_id, fecha, quincenas, total, adeudo, estado) 
                                          VALUES (?, ?, ?, ?, ?, ?)''',
                                       (cliente_id, str(fecha_vale), quincenas, total, adeudo, estado))
                        
                        venta_id = cursor.lastrowid
                        
                        # --- C. EXTRAER LOS PAGOS DE ESTA VENTA ---
                        fila_titulos_pagos = fila + 3
                        fila_fechas_pagos = fila + 4
                        fila_montos_pagos = fila + 5
                        
                        for col_pago in range(1, hoja.max_column + 1):
                            titulo_pago = hoja.cell(row=fila_titulos_pagos, column=col_pago).value
                            
                            if isinstance(titulo_pago, str) and titulo_pago.strip().upper().startswith("PAGO"):
                                fecha_pago = hoja.cell(row=fila_fechas_pagos, column=col_pago).value
                                monto_pago = hoja.cell(row=fila_montos_pagos, column=col_pago).value
                                
                                if monto_pago is not None:
                                    try:
                                        monto_float = float(monto_pago)
                                        if monto_float > 0:
                                            try:
                                                num_pago = int(titulo_pago.strip().upper().replace("PAGO", "").strip())
                                            except:
                                                num_pago = 0
                                                
                                            cursor.execute('''INSERT INTO pagos (venta_id, numero_pago, fecha, monto) 
                                                              VALUES (?, ?, ?, ?)''',
                                                           (venta_id, num_pago, str(fecha_pago), monto_float))
                                    except ValueError:
                                        pass

            print(f"[{i+1}/{total_archivos}] Migrado: {archivo} (Con abonos, sin folio)")

        except Exception as e:
            print(f"ERROR al procesar {archivo}: {e}")

    conn.commit()
    conn.close()
    print("\n¡Migración terminada con éxito!")

if __name__ == "__main__":
    migrar_archivos()